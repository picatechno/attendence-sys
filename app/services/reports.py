"""Report generation service - Attendance reports in XLSX and PDF formats"""
import io
import math
from datetime import date, datetime, timedelta, timezone
from typing import Optional

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_

from app.models.attendance import Attendance, Shift
from app.models.employee import Employee
from app.models.organization import Department
from app.models.organization import Organization


class ReportService:

    @staticmethod
    def _seconds_to_hours(seconds: Optional[int]) -> str:
        if not seconds:
            return "0h 0m"
        hours = seconds // 3600
        mins = (seconds % 3600) // 60
        return f"{hours}h {mins}m"

    @staticmethod
    async def get_daily_report(db: AsyncSession, org_id: str, report_date: date) -> list[dict]:
        """Get attendance report for all employees in an org on a given date."""
        result = await db.execute(
            select(Attendance).where(
                Attendance.date == report_date,
            ).order_by(Attendance.employee_id)
        )
        records = result.scalars().all()
        att_map = {r.employee_id: r for r in records}

        emp_result = await db.execute(
            select(Employee).where(
                Employee.org_id == org_id, Employee.is_active == True
            ).order_by(Employee.employee_code)
        )
        employees = emp_result.scalars().all()

        report = []
        for emp in employees:
            att = att_map.get(emp.id)
            report.append({
                "employee_code": emp.employee_code,
                "employee_name": f"{emp.first_name} {emp.last_name}",
                "department": emp.department.name if emp.department else "",
                "date": report_date.isoformat(),
                "clock_in": att.clock_in.isoformat() if att and att.clock_in else "",
                "clock_out": att.clock_out.isoformat() if att and att.clock_out else "",
                "status": att.status if att else "absent",
                "total_hours": ReportService._seconds_to_hours(att.total_work_hours if att else None),
                "net_hours": ReportService._seconds_to_hours(att.net_work_hours if att else None),
                "late_minutes": att.late_minutes if att else 0,
                "early_leave": att.early_leave_minutes if att else 0,
                "overtime": ReportService._seconds_to_hours(att.overtime_hours if att else None),
            })
        return report

    @staticmethod
    async def get_monthly_summary(db: AsyncSession, org_id: str, year: int, month: int, employee_id: Optional[str] = None) -> list[dict]:
        """Get monthly attendance summary."""
        q = select(Attendance).where(
            func.extract("year", Attendance.date) == year,
            func.extract("month", Attendance.date) == month,
        )
        if employee_id:
            q = q.where(Attendance.employee_id == employee_id)
        q = q.order_by(Attendance.employee_id, Attendance.date)
        result = await db.execute(q)
        records = result.scalars().all()

        emp_ids = list(set(r.employee_id for r in records)) if not employee_id else [employee_id]

        emp_q = select(Employee).where(Employee.org_id == org_id, Employee.is_active == True)
        if emp_ids:
            emp_q = emp_q.where(Employee.id.in_(emp_ids))
        emp_result = await db.execute(emp_q.order_by(Employee.employee_code))
        employees = emp_result.scalars().all()
        emp_map = {e.id: e for e in employees}

        summary_map = {}
        for att in records:
            if att.employee_id not in summary_map:
                emp = emp_map.get(att.employee_id)
                summary_map[att.employee_id] = {
                    "employee_code": emp.employee_code if emp else "",
                    "employee_name": f"{emp.first_name} {emp.last_name}" if emp else "",
                    "department": emp.department.name if emp and emp.department else "",
                    "present_days": 0,
                    "absent_days": 0,
                    "late_days": 0,
                    "holidays": 0,
                    "week_offs": 0,
                    "total_work_seconds": 0,
                    "total_overtime_seconds": 0,
                }
            s = summary_map[att.employee_id]
            if att.status == "present":
                s["present_days"] += 1
            elif att.status == "absent":
                s["absent_days"] += 1
            elif att.status == "late":
                s["late_days"] += 1
                s["present_days"] += 1
            elif att.status == "holiday":
                s["holidays"] += 1
            elif att.status == "week_off":
                s["week_offs"] += 1
            if att.net_work_hours:
                s["total_work_seconds"] += att.net_work_hours
            if att.overtime_hours:
                s["total_overtime_seconds"] += att.overtime_hours

        result_list = []
        for sid, s in summary_map.items():
            s["total_hours"] = ReportService._seconds_to_hours(s.pop("total_work_seconds"))
            s["total_overtime"] = ReportService._seconds_to_hours(s.pop("total_overtime_seconds"))
            result_list.append(s)
        return result_list

    @staticmethod
    async def get_employee_report(db: AsyncSession, employee_id: str, date_from: date, date_to: date) -> list[dict]:
        """Get detailed attendance report for a single employee in a date range."""
        result = await db.execute(
            select(Attendance).where(
                Attendance.employee_id == employee_id,
                Attendance.date >= date_from,
                Attendance.date <= date_to,
            ).order_by(Attendance.date)
        )
        records = result.scalars().all()

        emp_result = await db.execute(select(Employee).where(Employee.id == employee_id))
        emp = emp_result.scalar_one_or_none()

        report = []
        for att in records:
            report.append({
                "employee_code": emp.employee_code if emp else "",
                "employee_name": f"{emp.first_name} {emp.last_name}" if emp else "",
                "date": att.date.isoformat(),
                "clock_in": att.clock_in.isoformat() if att.clock_in else "",
                "clock_out": att.clock_out.isoformat() if att.clock_out else "",
                "status": att.status,
                "total_hours": ReportService._seconds_to_hours(att.total_work_hours),
                "net_hours": ReportService._seconds_to_hours(att.net_work_hours),
                "late_minutes": att.late_minutes or 0,
                "early_leave": att.early_leave_minutes or 0,
                "overtime": ReportService._seconds_to_hours(att.overtime_hours),
            })
        return report

    @staticmethod
    def generate_xlsx(report_data: list[dict], sheet_name: str = "Report") -> bytes:
        """Generate XLSX bytes from report data."""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        if not report_data:
            return b""

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = list(report_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_idx, row_data in enumerate(report_data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_idx in range(2, len(report_data) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_pdf(report_data: list[dict], title: str = "Attendance Report") -> bytes:
        """Generate PDF bytes from report data using ReportLab."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=20, leftMargin=20)
        styles = getSampleStyleSheet()

        elements = []
        elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
        elements.append(Spacer(1, 12))

        if not report_data:
            elements.append(Paragraph("No data available.", styles["Normal"]))
            doc.build(elements)
            return output.getvalue()

        headers = list(report_data[0].keys())
        display_headers = [h.replace("_", " ").title() for h in headers]

        data_rows = [[Paragraph(str(r.get(h, "")), styles["Normal"]) for h in headers] for r in report_data]
        table_data = [[Paragraph(h, styles["Heading6"]) for h in display_headers]] + data_rows

        col_width = max(50, int((doc.pagesize[0] - 40) / len(headers)))
        table = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#DCE6F1")]),
        ]))
        elements.append(table)
        doc.build(elements)
        return output.getvalue()
